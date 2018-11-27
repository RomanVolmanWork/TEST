import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

dicResults={"PPR":"","PPC":"","PPM":"","MEAN":"","MEDIAN":"","Percentile3DERROR":"",} 
PassPracent = 0.99
passFlag = True
resultString=""
arr = sys.argv[1].split('\\')
fileName = arr[len(arr) - 2]
fileNumber = str(int(sys.argv[3]) + 1)
revision =  sys.argv[2]
with open(sys.argv[1], 'r') as pointCloudCompareResult:
    for line in pointCloudCompareResult:
        arrData = line.split(":")

        if("ref pcl" in arrData[0]):
           dicResults["PPR"] = float(arrData[1].strip()) 
        elif ("compared pcl" in arrData[0]):
            dicResults["PPC"] = float(arrData[1].strip())
        elif ("matched number" in arrData[0]):
            dicResults["PPM"] = float(arrData[1].strip())
        elif ("mean" in arrData[0]):
            dicResults["MEAN"] = float(arrData[1].strip())
        elif ("median" in arrData[0]):
            dicResults["MEDIAN"] = float(arrData[1].strip())
        elif ("percentile" in arrData[0]):
            dicResults["Percentile3DERROR"] = float(arrData[1].strip())  


filepath = r"P:\Movidius\DataForMoviTest\testResultPointCloudCompare.xlsx"

# isFileExsistBoolean = os.path.exists(filepath)
# if isFileExsistBoolean:
book = load_workbook(filepath)
if not revision in book.sheetnames:
    print('revision not exists')
    sheetName = book['Sheet']
    sheetName.title = revision
    sheetName['B1'].value = 'PPM/PPC'
    sheetName['C1'].value = 'PPM/PPR'
    sheetName['D1'].value = 'MEDIAN'
    sheetName['E1'].value = 'MEAN'
    sheetName['F1'].value = 'Percentile 3D ERROR'
    sheetName['G1'].value = 'File Name'
    sheetName['A1'].value = 'Result'
else:
    sheetName = book[revision]

A = "A"+ fileNumber
B = "B"+ fileNumber
C = "C"+ fileNumber
D = "D"+ fileNumber
E = "E"+ fileNumber
F = "F"+ fileNumber
G = "G"+ fileNumber

if not(dicResults["PPM"]/dicResults["PPC"] > PassPracent): 
   passFlag=False

if not(dicResults["PPM"]/dicResults["PPR"] > PassPracent):
   passFlag=False

if not(dicResults["MEDIAN"] < 1):
   passFlag=False

if not(dicResults["MEAN"] < 1):
   passFlag=False

if not(dicResults["Percentile3DERROR"] < 2):
   passFlag=False

if passFlag:
   resultString = "Pass"
   green = Font(color=colors.GREEN)
   sheetName[A].font = green
   print (resultString)
else:
   resultString = "Fail"
   red = Font(color=colors.RED)
   sheetName[A].font = red

sheetName[B].value = dicResults["PPM"]/dicResults["PPC"]
sheetName[C].value = dicResults["PPM"]/dicResults["PPR"]
sheetName[D].value = dicResults["MEDIAN"]
sheetName[E].value = dicResults["MEAN"]
sheetName[F].value = dicResults["Percentile3DERROR"]
sheetName[G].value = fileName
sheetName[A].value = resultString

book.save(filepath)

    











